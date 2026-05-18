import json
from pathlib import Path
from openpyxl import load_workbook
from rrl.db import connect, init_schema
from rrl.output.matrix import write_matrix, MATRIX_COLUMNS

def _seed(conn, pid, tier, **kw):
    base = dict(
        paper_id=pid, title="Title", authors_json=json.dumps([{"family":"S","given":"J"}]),
        year=2023, era_tag="post_chatgpt", venue="V", publisher="Springer",
        work_type="journal-article", doi=f"10.1/{pid}", language="en",
        is_in_doaj=1, is_peer_reviewed=1, is_oa=1, oa_status="gold",
        citation_count=5, topic_match_score=3.0,
        included=1, quality_tier=tier,
        pdf_filename="2023/x.pdf", pdf_status="downloaded",
        first_seen_at="now", last_updated_at="now",
    )
    base.update(kw)
    keys = ",".join(base.keys()); qs = ",".join(["?"] * len(base))
    conn.execute(f"INSERT INTO papers ({keys}) VALUES ({qs})", tuple(base.values()))

def test_matrix_has_two_sheets_and_expected_columns(tmp_path):
    conn = connect(tmp_path / "rrl.sqlite"); init_schema(conn)
    _seed(conn, "p1", "high_confidence")
    _seed(conn, "p2", "review_needed", title="Other", pdf_filename="2023/y.pdf")
    out = tmp_path / "out/matrix.xlsx"
    write_matrix(conn, out)
    wb = load_workbook(out)
    assert "high_confidence" in wb.sheetnames
    assert "review_needed" in wb.sheetnames
    hc = wb["high_confidence"]
    headers = [c.value for c in hc[1]]
    assert headers == MATRIX_COLUMNS
    assert "pdf_status" in headers
    assert hc.cell(row=2, column=1).value == "p1"
    rn = wb["review_needed"]
    assert rn.cell(row=2, column=1).value == "p2"

def test_matrix_excludes_unincluded_and_merged_but_includes_unretrievable(tmp_path):
    """After the OA pivot the matrix includes EVERY included, non-merged paper
    regardless of pdf_status. The pdf_status column tells the reader whether
    the file is downloaded / not_retrievable."""
    conn = connect(tmp_path / "rrl.sqlite"); init_schema(conn)
    _seed(conn, "p1", "high_confidence")
    _seed(conn, "p_excluded", "high_confidence", included=0)
    _seed(conn, "p_not_retrievable", "high_confidence",
          pdf_status="not_retrievable", pdf_filename=None)
    _seed(conn, "p_merged", "high_confidence")
    conn.execute("INSERT INTO paper_merges (loser_id, winner_id, merged_at, merged_by) "
                 "VALUES ('p_merged','p1','now','manual')")
    out = tmp_path / "out/matrix.xlsx"
    write_matrix(conn, out)
    wb = load_workbook(out)
    hc = wb["high_confidence"]
    ids = [hc.cell(row=i, column=1).value for i in range(2, hc.max_row + 1)]
    # p1 (downloaded) AND p_not_retrievable both appear; excluded + merged excluded.
    assert set(ids) == {"p1", "p_not_retrievable"}
