"""End-to-end smoke test: every stage runs against canned HTTP responses,
producing a real xlsx + manifest + README appendix on disk."""
import json
import responses
from openpyxl import load_workbook
from click.testing import CliRunner
from pathlib import Path
from rrl.cli import main
from rrl.output.readme import BEGIN_MARK, END_MARK

@responses.activate
def test_full_pipeline_smoke(tmp_path, monkeypatch, fixtures_dir):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("OPENALEX_EMAIL", "t@e.com")
    monkeypatch.delenv("ELSEVIER_API_KEY", raising=False)
    monkeypatch.delenv("ELSEVIER_INSTTOKEN", raising=False)
    monkeypatch.delenv("CORE_API_KEY", raising=False)
    monkeypatch.delenv("SEMANTIC_SCHOLAR_API_KEY", raising=False)

    oa1 = json.loads((fixtures_dir / "openalex_page1.json").read_text())
    oa2 = json.loads((fixtures_dir / "openalex_page2.json").read_text())
    eric = json.loads((fixtures_dir / "eric_response.json").read_text())
    eric_empty = {"response": {"numFound": 2, "start": 2000, "docs": []}}
    s2 = json.loads((fixtures_dir / "s2_bulk_response.json").read_text())
    s2_empty = {"token": None, "data": []}
    responses.add(responses.GET, "https://api.openalex.org/works", json=oa1)
    responses.add(responses.GET, "https://api.openalex.org/works", json=oa2)
    responses.add(responses.GET, "https://api.ies.ed.gov/eric/", json=eric)
    responses.add(responses.GET, "https://api.ies.ed.gov/eric/", json=eric_empty)
    responses.add(responses.GET, "https://api.semanticscholar.org/graph/v1/paper/search/bulk", json=s2)
    responses.add(responses.GET, "https://api.semanticscholar.org/graph/v1/paper/search/bulk", json=s2_empty)

    # Enrich: DOAJ + Unpaywall.
    responses.add(responses.GET, "https://doaj.org/api/v3/search/journals/issn:1234-5678",
                  json={"results": [{"id": "abc"}]}, status=200)
    for doi in ("10.1/aaa", "10.1/ccc", "10.1/zzz"):
        responses.add(responses.GET, f"https://api.unpaywall.org/v2/{doi}",
                      json={"best_oa_location": {"url_for_pdf": f"https://x/{doi.replace('/','_')}.pdf"}},
                      status=200)

    # PDFs: return valid bytes.
    pdf_bytes = (fixtures_dir / "sample.pdf").read_bytes()
    for doi in ("10.1/aaa", "10.1/ccc", "10.1/zzz"):
        responses.add(responses.GET, f"https://x/{doi.replace('/','_')}.pdf",
                      body=pdf_bytes, content_type="application/pdf", status=200)
    # OpenAlex's own oa pdf for W111/W333 (fallback)
    responses.add(responses.GET, "https://example.com/a.pdf", body=pdf_bytes,
                  content_type="application/pdf", status=200)
    responses.add(responses.GET, "https://example.com/c.pdf", body=pdf_bytes,
                  content_type="application/pdf", status=200)

    # README with markers must exist.
    (tmp_path / "README.md").write_text(
        f"# RRL\n\nIntro.\n\n{BEGIN_MARK}\nplaceholder\n{END_MARK}\n", encoding="utf-8")

    runner = CliRunner()
    r = runner.invoke(main, ["all"])
    assert r.exit_code == 0, r.output

    matrix = tmp_path / "output/rrl_matrix.xlsx"
    manifest = tmp_path / "output/run_manifest.json"
    assert matrix.exists() and manifest.exists()
    wb = load_workbook(matrix)
    assert {"high_confidence", "review_needed"} <= set(wb.sheetnames)
    rows = sum(s.max_row - 1 for s in (wb["high_confidence"], wb["review_needed"]))
    assert rows >= 1
    readme = (tmp_path / "README.md").read_text(encoding="utf-8")
    assert "Run statistics" in readme
    pdfs = list((tmp_path / "pdfs").rglob("*.pdf"))
    assert len(pdfs) >= 1
    m = json.loads(manifest.read_text())
    assert m["pipeline_version"]
    assert "counts" in m and "matrix_sha256" in m
