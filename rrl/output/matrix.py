"""Write the two-sheet xlsx matrix from the papers table."""
from __future__ import annotations
import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

MATRIX_COLUMNS = [
    "paper_id", "title", "authors", "year", "era_tag", "venue", "publisher",
    "work_type", "doi", "language", "is_in_doaj", "is_peer_reviewed",
    "is_oa", "oa_status", "citation_count", "topic_match_score",
    "pdf_filename", "source_apis", "abstract",
]

def _yes_no_na(v):
    if v is None:
        return "N/A"
    return "Yes" if v else "No"

def _authors_str(authors_json: str) -> str:
    try:
        authors = json.loads(authors_json or "[]")
    except Exception:
        return ""
    parts = []
    for a in authors:
        family = a.get("family") or ""
        given = (a.get("given") or "").strip()
        initial = f", {given[0]}." if given else ""
        parts.append(f"{family}{initial}".strip(", "))
    return "; ".join(parts)

def _source_apis(conn: sqlite3.Connection, paper_id: str) -> str:
    rows = conn.execute(
        """SELECT DISTINCT rr.adapter FROM raw_records rr
           JOIN paper_sources ps ON ps.raw_id = rr.raw_id
           WHERE ps.paper_id = ?""",
        (paper_id,),
    ).fetchall()
    return ",".join(sorted({r["adapter"] for r in rows}))

def _row_values(conn, p) -> list:
    return [
        p["paper_id"], p["title"], _authors_str(p["authors_json"]),
        p["year"], p["era_tag"], p["venue"], p["publisher"],
        p["work_type"], p["doi"], p["language"],
        _yes_no_na(p["is_in_doaj"]), _yes_no_na(p["is_peer_reviewed"]),
        _yes_no_na(p["is_oa"]), p["oa_status"],
        p["citation_count"], p["topic_match_score"],
        p["pdf_filename"], _source_apis(conn, p["paper_id"]), p["abstract"],
    ]

QUERY = """
SELECT * FROM papers
WHERE included = 1
  AND pdf_status = 'downloaded'
  AND paper_id NOT IN (SELECT loser_id FROM paper_merges)
  AND quality_tier = ?
ORDER BY year DESC, title
"""

def _write_sheet(ws, conn, tier: str) -> None:
    ws.append(MATRIX_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for p in conn.execute(QUERY, (tier,)).fetchall():
        ws.append(_row_values(conn, p))
        if p["doi"]:
            ws.cell(row=ws.max_row, column=MATRIX_COLUMNS.index("doi") + 1).hyperlink = f"https://doi.org/{p['doi']}"
        if p["pdf_filename"]:
            ws.cell(row=ws.max_row, column=MATRIX_COLUMNS.index("pdf_filename") + 1).hyperlink = f"pdfs/{p['pdf_filename']}"
    for col_idx, name in enumerate(MATRIX_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if name == "abstract":
            ws.column_dimensions[letter].width = 60
            for c in ws[letter][1:]:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            max_len = max((len(str(c.value or "")) for c in ws[letter]), default=10)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)

def write_matrix(conn: sqlite3.Connection, out_path: Path) -> dict:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    hc = wb.create_sheet("high_confidence")
    rn = wb.create_sheet("review_needed")
    _write_sheet(hc, conn, "high_confidence")
    _write_sheet(rn, conn, "review_needed")
    wb.save(out_path)
    return {"high_confidence": hc.max_row - 1, "review_needed": rn.max_row - 1}
